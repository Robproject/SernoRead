import tkinter
from tkinter import filedialog
import cv2
import pytesseract
import os
import numpy as np
import openpyxl

# select paths
def get_file():
    global xlsx_path
    xlsx_path = filedialog.askopenfilename()

def get_dir():
    global dir_path
    dir_path = filedialog.askdirectory(title='Select Folder')

def get_outline(image):
    contours,hierarchy = cv2.findContours(image,cv2.RETR_EXTERNAL,cv2.CHAIN_APPROX_SIMPLE)
    # only use main contour
    threshold_area = 3000000
    for cnt in contours:
        area = cv2.contourArea(cnt)
        if area > threshold_area:
            x, y, w, h = cv2.boundingRect(cnt)
            return x,y,w,h

def fix_orientation(img):
    # orient vertically if not
    if img.shape[1] > img.shape[0]:
        img = cv2.rotate(img, cv2.ROTATE_90_CLOCKWISE)
    # crop spot and get avg color
    crop = img[1675:1825, 140:310]
    average = np.average(crop)
    # rotate 180 if flipped
    if average > 100:
        img = cv2.rotate(img, cv2.ROTATE_180)
    return img

def process_photo(img_path):
    img = cv2.imread(img_path,0)
    # process img then crop board and orient
    img = cv2.medianBlur(img, 5)
    thresh = cv2.threshold(img, 100, 255, cv2.THRESH_BINARY_INV)[1]
    x,y,w,h = get_outline(thresh)
    crop = img[y:y+h,x:x+w]
    img = fix_orientation(crop)
    return img

def get_sernos():
    serno_dict = {}
    for filename in os.listdir(dir_path):
        f = os.path.join(dir_path, filename)
        # check if file
        if os.path.isfile(f):
            if "INPUT" in os.path.basename(f):
                img = process_photo(f)
                # crop serno area, thresh, scale
                img = img[70:140, 875:1020]
                img = cv2.threshold(img, 110, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)[1]
                img = cv2.resize(img, (0, 0), fx=5, fy=5)
                # get text
                serno = pytesseract.image_to_string(img, config="-c tessedit_char_whitelist=0123456789 --psm 8") 
                filenumberjpg = os.path.basename(f).split("_")[3]
                filenumber = filenumberjpg.split(".")[0]
                serno_dict[int(filenumber)] = serno[0:5]
    write_dict, update_dict, wb_obj = compare_data(serno_dict)
    sheet = wb_obj['Pictures']
    for key in update_dict:
        for row in sheet.iter_rows(1, sheet.max_row):
            if row[0].value == int(key):
                row[1].value = update_dict[key]
    for key in write_dict:
        sheet.cell(column = 1, row = sheet.max_row + 1).value = int(key)
        sheet.cell(column = 2, row = sheet.max_row).value = int(write_dict[key])
    wb_obj.save(xlsx_path)
    print("Complete")

def get_xlsxdata():
    wb_obj = openpyxl.load_workbook(xlsx_path)
    sheet = wb_obj['Pictures']
    xlsx_dict = {}
    for row in sheet.iter_rows(1, sheet.max_row):
        xlsx_dict[row[0].value] = row[1].value
    return xlsx_dict, wb_obj

def compare_data(serno_dict):
    xlsx_dict, wb_obj = get_xlsxdata()
    sheet = wb_obj['Pictures']
    write_dict = {}
    update_dict = {}
    for key in serno_dict:
        if key in xlsx_dict:
            if xlsx_dict[key] == int(serno_dict[key]):
                print(str(key) + " is already populated and is correct.")
            elif xlsx_dict[key] == None:
                update_dict[key] = int(serno_dict[key])
            else:
                print("xlsx serno: " + str(xlsx_dict[key]) + " doesn't match ocr serno: " + str(serno_dict[key]))
        else:
            write_dict[key] = serno_dict[key]
    return write_dict, update_dict, wb_obj

if __name__ == '__main__':
    rootwindow = tkinter.Tk()
    rootwindow.title("WLM300 SerNo OCR")
    rootwindow.geometry("350x275")

    xlsx_path = '/'
    dir_path = '/'

    button_getfile = tkinter.Button(rootwindow, text="select log.xlsx", command = get_file)
    button_getdir = tkinter.Button(rootwindow, text ="select photo dir", command = get_dir)
    button_processphotos = tkinter.Button(rootwindow, text="Copy SerNos", command = lambda: get_sernos())
    #button_printxlsx = tkinter.Button(rootwindow, text="Print XLSX", command = lambda: get_xlsxdata(xlsx_path))


    button_getfile.grid(row=0,column=0,padx=10,pady=10)
    button_getdir.grid(row=0,column=1,padx=10,pady=10)
    button_processphotos.grid(row=1,column=0,padx=10,pady=10)
    #button_printxlsx.grid(row=1,column=1)

    rootwindow.mainloop()
    
